#!/usr/bin/env python3
"""
Compliance Rule Batch Runner
============================
Reads compliance rules from an Excel file (Rules sheet), executes each rule
against portfolio.db, and writes a structured results report to Excel.

Skill: compliance-batch-runner
Depends on: compliance-rule-generator (code generation done at authoring time)

Usage:
    python scripts/compliance_batch_runner.py --fund-id 1 --date 2024-01-31
    python scripts/compliance_batch_runner.py --fund-id 1 --date 2024-01-31 \
        --input scripts/my_rules.xlsx --output results.xlsx
"""

import argparse
import os
import sqlite3
import traceback
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DB_PATH = os.path.join(PROJECT_ROOT, "portfolio.db")
DEFAULT_RULES_PATH = os.path.join(SCRIPT_DIR, "compliance_rules.xlsx")

# ---------------------------------------------------------------------------
# execute_sql — called by generated functions at runtime
# ---------------------------------------------------------------------------

def execute_sql(sql: str) -> pd.DataFrame:
    """Execute SQL against portfolio.db and return a DataFrame."""
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql_query(sql, conn)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Generated compliance functions
# (SQL authored at generation time via compliance-rule-generator skill,
#  adapted to the portfolio.db schema: funds + positions tables.
#  NAV basis: funds.aum_usd (millions); weight basis: positions.weight_pct)
# ---------------------------------------------------------------------------

def check_max_equity_exposure(fund_id: str, as_of_date: str, threshold: float = 60.0) -> dict:
    """
    Rule: max 60% of portfolio in Equity
    Rule type: MAX_PCT_NAV
    Checks that total Equity weight does not exceed threshold%.

    Term resolution: "Equity" = positions with asset_class = 'Equity'
    Term resolution: "% of NAV" = SUM(weight_pct) for the asset class
    NOTE: as_of_date not stored in schema; all positions are current snapshot.
    """
    rule_text = "max 60% of portfolio in Equity"
    rule_type = "MAX_PCT_NAV"
    try:
        # Query 1: Total Equity weight for the fund
        # text2sql: "Sum of weight_pct for Equity positions belonging to a given fund"
        sql_equity = f"""
            SELECT SUM(weight_pct) AS equity_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND asset_class = 'Equity'
        """
        df = execute_sql(sql_equity)
        equity_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0

        passed = equity_weight <= threshold

        breaches = []
        if not passed:
            # Query 2: Position-level breakdown of Equity holdings
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND asset_class = 'Equity'
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(equity_weight, 4),
            "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Equity exposure is {equity_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_no_restricted_positions(fund_id: str, as_of_date: str, threshold: int = 0) -> dict:
    """
    Rule: no positions with Restricted compliance status
    Rule type: PROHIBITED
    """
    rule_text = "no positions with Restricted compliance status"
    rule_type = "PROHIBITED"
    try:
        # text2sql: "Count of positions with compliance_status = Restricted for a given fund"
        sql_count = f"""
            SELECT COUNT(*) AS restricted_count
            FROM positions
            WHERE fund_id = {fund_id}
              AND compliance_status = 'Restricted'
        """
        df = execute_sql(sql_count)
        count = int(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0
        passed = count == 0

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND compliance_status = 'Restricted'
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": count, "threshold": threshold, "unit": "count",
            "breaches": breaches,
            "message": (
                "No Restricted positions found." if passed
                else f"{count} Restricted position(s) found — rule breached."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "count", e)


def check_max_single_country_concentration(fund_id: str, as_of_date: str, threshold: float = 30.0) -> dict:
    """
    Rule: max 30% of portfolio in any single country
    Rule type: CONCENTRATION
    """
    rule_text = "max 30% of portfolio in any single country"
    rule_type = "CONCENTRATION"
    try:
        # text2sql: "Sum of weight_pct grouped by country for a given fund, ordered descending"
        sql = f"""
            SELECT country, SUM(weight_pct) AS country_weight
            FROM positions
            WHERE fund_id = {fund_id}
            GROUP BY country
            ORDER BY country_weight DESC
        """
        df = execute_sql(sql)
        if df.empty:
            return _pass(rule_text, rule_type, fund_id, as_of_date, 0.0, threshold, "% of NAV",
                         "No positions found.")

        max_weight = float(df.iloc[0]["country_weight"])
        passed = max_weight <= threshold

        breaches = []
        if not passed:
            breaching = df[df["country_weight"] > threshold]
            for _, row in breaching.iterrows():
                breaches.append({
                    "identifier": str(row["country"]),
                    "description": f"Country: {row['country']}",
                    "value": float(row["country_weight"]),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(max_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Max country concentration is {max_weight:.2f}% ({df.iloc[0]['country']}), "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_single_sector_concentration(fund_id: str, as_of_date: str, threshold: float = 30.0) -> dict:
    """
    Rule: max 30% of portfolio in any single sector
    Rule type: CONCENTRATION
    """
    rule_text = "max 30% of portfolio in any single sector"
    rule_type = "CONCENTRATION"
    try:
        # text2sql: "Sum of weight_pct grouped by sector for a given fund, ordered descending"
        sql = f"""
            SELECT sector, SUM(weight_pct) AS sector_weight
            FROM positions
            WHERE fund_id = {fund_id}
            GROUP BY sector
            ORDER BY sector_weight DESC
        """
        df = execute_sql(sql)
        if df.empty:
            return _pass(rule_text, rule_type, fund_id, as_of_date, 0.0, threshold, "% of NAV",
                         "No positions found.")

        max_weight = float(df.iloc[0]["sector_weight"])
        passed = max_weight <= threshold

        breaches = []
        if not passed:
            breaching = df[df["sector_weight"] > threshold]
            for _, row in breaching.iterrows():
                breaches.append({
                    "identifier": str(row["sector"]),
                    "description": f"Sector: {row['sector']}",
                    "value": float(row["sector_weight"]),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(max_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Max sector concentration is {max_weight:.2f}% ({df.iloc[0]['sector']}), "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_bond_exposure(fund_id: str, as_of_date: str, threshold: float = 40.0) -> dict:
    """
    Rule: max 40% of portfolio in Bonds
    Rule type: MAX_PCT_NAV
    Term resolution: "Bonds" = positions with asset_class = 'Bond'
    """
    rule_text = "max 40% of portfolio in Bonds"
    rule_type = "MAX_PCT_NAV"
    try:
        sql = f"""
            SELECT SUM(weight_pct) AS bond_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND asset_class = 'Bond'
        """
        df = execute_sql(sql)
        bond_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = bond_weight <= threshold

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND asset_class = 'Bond'
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(bond_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Bond exposure is {bond_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_min_cash_exposure(fund_id: str, as_of_date: str, threshold: float = 2.0) -> dict:
    """
    Rule: min 2% of portfolio in Cash
    Rule type: MIN_PCT_NAV
    Term resolution: "Cash" = positions with asset_class = 'Cash'
    """
    rule_text = "min 2% of portfolio in Cash"
    rule_type = "MIN_PCT_NAV"
    try:
        sql = f"""
            SELECT SUM(weight_pct) AS cash_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND asset_class = 'Cash'
        """
        df = execute_sql(sql)
        cash_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = cash_weight >= threshold

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(cash_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": [],
            "message": (
                f"Cash is {cash_weight:.2f}% of portfolio, "
                f"{'meets' if passed else 'below'} minimum {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_single_position_weight(fund_id: str, as_of_date: str, threshold: float = 15.0) -> dict:
    """
    Rule: no single position to exceed 15% of NAV
    Rule type: CONCENTRATION
    """
    rule_text = "no single position to exceed 15% of NAV"
    rule_type = "CONCENTRATION"
    try:
        # text2sql: "All positions for a given fund with their weight_pct, ordered by weight descending"
        sql = f"""
            SELECT isin AS identifier, security_name AS description, weight_pct AS value
            FROM positions
            WHERE fund_id = {fund_id}
            ORDER BY weight_pct DESC
        """
        df = execute_sql(sql)
        if df.empty:
            return _pass(rule_text, rule_type, fund_id, as_of_date, 0.0, threshold, "% of NAV",
                         "No positions found.")

        max_weight = float(df.iloc[0]["value"])
        passed = max_weight <= threshold

        breaches = []
        if not passed:
            breaching = df[df["value"] > threshold]
            for _, row in breaching.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(max_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Largest position is {max_weight:.2f}% ({df.iloc[0]['description']}), "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_review_status_exposure(fund_id: str, as_of_date: str, threshold: float = 20.0) -> dict:
    """
    Rule: max 20% of portfolio in Review status positions
    Rule type: MAX_PCT_NAV
    """
    rule_text = "max 20% of portfolio in Review status positions"
    rule_type = "MAX_PCT_NAV"
    try:
        sql = f"""
            SELECT SUM(weight_pct) AS review_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND compliance_status = 'Review'
        """
        df = execute_sql(sql)
        review_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = review_weight <= threshold

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND compliance_status = 'Review'
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(review_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Review-status positions are {review_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_commodity_exposure(fund_id: str, as_of_date: str, threshold: float = 15.0) -> dict:
    """
    Rule: max 15% of portfolio in Commodity
    Rule type: MAX_PCT_NAV
    Term resolution: "Commodity" = positions with asset_class = 'Commodity'
    """
    rule_text = "max 15% of portfolio in Commodity"
    rule_type = "MAX_PCT_NAV"
    try:
        sql = f"""
            SELECT SUM(weight_pct) AS commodity_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND asset_class = 'Commodity'
        """
        df = execute_sql(sql)
        commodity_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = commodity_weight <= threshold

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND asset_class = 'Commodity'
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(commodity_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Commodity exposure is {commodity_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_max_equity_etf_combined(fund_id: str, as_of_date: str, threshold: float = 50.0) -> dict:
    """
    Rule: max 50% of portfolio in Equity plus ETF combined
    Rule type: MAX_PCT_NAV
    Term resolution: "Equity plus ETF" = asset_class IN ('Equity', 'ETF')
    """
    rule_text = "max 50% of portfolio in Equity plus ETF combined"
    rule_type = "MAX_PCT_NAV"
    try:
        sql = f"""
            SELECT SUM(weight_pct) AS equity_etf_weight
            FROM positions
            WHERE fund_id = {fund_id}
              AND asset_class IN ('Equity', 'ETF')
        """
        df = execute_sql(sql)
        combined_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = combined_weight <= threshold

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                  AND asset_class IN ('Equity', 'ETF')
                ORDER BY weight_pct DESC
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(combined_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Equity+ETF combined is {combined_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


def check_min_position_count(fund_id: str, as_of_date: str, threshold: int = 10) -> dict:
    """
    Rule: minimum 10 positions in portfolio
    Rule type: MAX_COUNT (inverted — minimum count)
    """
    rule_text = "minimum 10 positions in portfolio"
    rule_type = "MAX_COUNT"
    try:
        sql = f"""
            SELECT COUNT(*) AS position_count
            FROM positions
            WHERE fund_id = {fund_id}
        """
        df = execute_sql(sql)
        count = int(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0
        passed = count >= threshold

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": count, "threshold": threshold, "unit": "count",
            "breaches": [],
            "message": (
                f"Portfolio has {count} positions, "
                f"{'meets' if passed else 'below'} minimum of {threshold}."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "count", e)


def check_top5_holdings_concentration(fund_id: str, as_of_date: str, threshold: float = 80.0) -> dict:
    """
    Rule: top 5 holdings max 80% of portfolio
    Rule type: CONCENTRATION
    """
    rule_text = "top 5 holdings max 80% of portfolio"
    rule_type = "CONCENTRATION"
    try:
        # text2sql: "Sum of weight_pct of the 5 largest positions for a given fund"
        sql = f"""
            SELECT SUM(weight_pct) AS top5_weight
            FROM (
                SELECT weight_pct
                FROM positions
                WHERE fund_id = {fund_id}
                ORDER BY weight_pct DESC
                LIMIT 5
            )
        """
        df = execute_sql(sql)
        top5_weight = float(df.iloc[0, 0]) if not df.empty and df.iloc[0, 0] is not None else 0.0
        passed = top5_weight <= threshold

        breaches = []
        if not passed:
            sql_breach = f"""
                SELECT isin AS identifier, security_name AS description, weight_pct AS value
                FROM positions
                WHERE fund_id = {fund_id}
                ORDER BY weight_pct DESC
                LIMIT 5
            """
            df_b = execute_sql(sql_breach)
            for _, row in df_b.iterrows():
                breaches.append({
                    "identifier": str(row.get("identifier") or ""),
                    "description": str(row.get("description") or ""),
                    "value": float(row.get("value") or 0),
                })

        return {
            "rule": rule_text, "rule_type": rule_type,
            "fund_id": fund_id, "as_of_date": as_of_date,
            "passed": passed,
            "metric_value": round(top5_weight, 4), "threshold": threshold, "unit": "% of NAV",
            "breaches": breaches,
            "message": (
                f"Top 5 holdings represent {top5_weight:.2f}% of portfolio, "
                f"{'within' if passed else 'exceeds'} max {threshold}%."
            ),
        }
    except Exception as e:
        return _error(rule_text, rule_type, fund_id, as_of_date, threshold, "% of NAV", e)


# ---------------------------------------------------------------------------
# Registry — maps rule_id to the generated function
# ---------------------------------------------------------------------------

RULE_REGISTRY = {
    "R001": check_max_equity_exposure,
    "R002": check_no_restricted_positions,
    "R003": check_max_single_country_concentration,
    "R004": check_max_single_sector_concentration,
    "R005": check_max_bond_exposure,
    "R006": check_min_cash_exposure,
    "R007": check_max_single_position_weight,
    "R008": check_max_review_status_exposure,
    "R009": check_max_commodity_exposure,
    "R010": check_max_equity_etf_combined,
    "R011": check_min_position_count,
    "R012": check_top5_holdings_concentration,
}

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _error(rule_text, rule_type, fund_id, as_of_date, threshold, unit, exc):
    return {
        "rule": rule_text, "rule_type": rule_type,
        "fund_id": fund_id, "as_of_date": as_of_date,
        "passed": False, "metric_value": None,
        "threshold": threshold, "unit": unit, "breaches": [],
        "message": f"Error during rule check: {exc}",
        "error": True,
    }

def _pass(rule_text, rule_type, fund_id, as_of_date, metric_value, threshold, unit, message):
    return {
        "rule": rule_text, "rule_type": rule_type,
        "fund_id": fund_id, "as_of_date": as_of_date,
        "passed": True, "metric_value": metric_value,
        "threshold": threshold, "unit": unit, "breaches": [],
        "message": message,
    }


# ---------------------------------------------------------------------------
# Step 1 — Read input Excel
# ---------------------------------------------------------------------------

def load_rules(input_path: str) -> list[dict]:
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input file not found: {input_path}")
    df = pd.read_excel(input_path, sheet_name="Rules", dtype=str)
    if "Rules" not in pd.ExcelFile(input_path).sheet_names:
        raise ValueError("Input file must contain a sheet named 'Rules'")
    df["enabled"] = df.get("enabled", pd.Series(["TRUE"] * len(df))).fillna("TRUE").str.upper()
    return df.to_dict("records")


# ---------------------------------------------------------------------------
# Step 2+3 — Resolve and execute each rule
# ---------------------------------------------------------------------------

def run_rules(rules: list[dict], fund_id: str, as_of_date: str) -> tuple[list[dict], list[dict]]:
    summary_rows = []
    breach_rows = []

    for rule in rules:
        rule_id = str(rule.get("rule_id", "")).strip()
        enabled = str(rule.get("enabled", "TRUE")).strip().upper()

        if enabled != "TRUE":
            summary_rows.append({
                "rule_id": rule_id,
                "rule_text": rule.get("rule_text", ""),
                "category": rule.get("category", ""),
                "status": "SKIPPED",
                "metric_value": "",
                "threshold": "",
                "unit": "",
                "message": "Rule disabled",
                "breach_count": 0,
                "notes": rule.get("notes", ""),
            })
            continue

        func = RULE_REGISTRY.get(rule_id)
        if func is None:
            print(f"  [WARN] No generated function for rule_id '{rule_id}' — recording ERROR")
            result = {
                "passed": False, "error": True,
                "metric_value": None, "threshold": None, "unit": "",
                "breaches": [],
                "message": f"No compliance function registered for rule_id '{rule_id}'",
            }
        else:
            kwargs = {"fund_id": fund_id, "as_of_date": as_of_date}
            override = rule.get("threshold_override", "")
            if override and str(override).strip() not in ("", "nan", "None"):
                try:
                    kwargs["threshold"] = float(override)
                except ValueError:
                    print(f"  [WARN] Non-numeric threshold_override '{override}' for {rule_id} — using default")
            try:
                result = func(**kwargs)
            except Exception as exc:
                result = {
                    "passed": False, "error": True,
                    "metric_value": None, "threshold": None, "unit": "",
                    "breaches": [],
                    "message": f"Uncaught exception: {exc}\n{traceback.format_exc()}",
                }

        status = "ERROR" if result.get("error") else ("PASS" if result["passed"] else "FAIL")
        summary_rows.append({
            "rule_id": rule_id,
            "rule_text": rule.get("rule_text", result.get("rule", "")),
            "category": rule.get("category", ""),
            "status": status,
            "metric_value": result.get("metric_value"),
            "threshold": result.get("threshold"),
            "unit": result.get("unit", ""),
            "message": result.get("message", ""),
            "breach_count": len(result.get("breaches", [])),
            "notes": rule.get("notes", ""),
        })

        for breach in result.get("breaches", []):
            breach_rows.append({
                "rule_id": rule_id,
                "rule_text": rule.get("rule_text", result.get("rule", "")),
                "identifier": breach.get("identifier", ""),
                "description": breach.get("description", ""),
                "value": breach.get("value"),
                "unit": result.get("unit", ""),
            })

    return summary_rows, breach_rows


# ---------------------------------------------------------------------------
# Step 5 — Write output Excel
# ---------------------------------------------------------------------------

def write_output(output_path: str, summary_rows: list[dict], breach_rows: list[dict]):
    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"

    headers = ["rule_id", "rule_text", "category", "status",
               "metric_value", "threshold", "unit", "message", "breach_count", "notes"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    status_fills = {
        "PASS":    PatternFill("solid", fgColor="C6EFCE"),
        "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
        "ERROR":   PatternFill("solid", fgColor="FFEB9C"),
        "SKIPPED": PatternFill("solid", fgColor="D9D9D9"),
    }

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    status_col_idx = headers.index("status") + 1
    for row_idx, row in enumerate(summary_rows, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row.get(h, ""))
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        status_cell.fill = status_fills.get(str(row.get("status", "")), PatternFill())

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # ---- Breaches sheet ----
    ws_b = wb.create_sheet("Breaches")
    b_headers = ["rule_id", "rule_text", "identifier", "description", "value", "unit"]
    breach_header_fill = PatternFill("solid", fgColor="C00000")
    breach_header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(b_headers, 1):
        cell = ws_b.cell(row=1, column=col, value=h)
        cell.fill = breach_header_fill
        cell.font = breach_header_font

    zebra_fills = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="FFE7E7")]
    current_rule = None
    zebra_idx = 0
    for row_idx, row in enumerate(breach_rows, 2):
        if row["rule_id"] != current_rule:
            current_rule = row["rule_id"]
            zebra_idx = 1 - zebra_idx
        fill = zebra_fills[zebra_idx]
        for col, h in enumerate(b_headers, 1):
            cell = ws_b.cell(row=row_idx, column=col, value=row.get(h, ""))
            cell.fill = fill

    ws_b.freeze_panes = "A2"
    for col in ws_b.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws_b.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Step 6 — Console summary
# ---------------------------------------------------------------------------

def print_summary(fund_id: str, as_of_date: str, summary_rows: list[dict], output_path: str):
    total = len(summary_rows)
    counts = {"PASS": 0, "FAIL": 0, "ERROR": 0, "SKIPPED": 0}
    for row in summary_rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    total_breaches = sum(int(row.get("breach_count") or 0) for row in summary_rows)
    enabled = total - counts["SKIPPED"]

    print("\n" + "=" * 60)
    print("Compliance Batch Run Complete")
    print(f"Fund:        {fund_id}")
    print(f"As of date:  {as_of_date}")
    print(f"Rules run:   {total}  ({enabled} enabled, {counts['SKIPPED']} skipped)")
    print("-" * 60)
    print(f"PASS:        {counts['PASS']}")
    print(f"FAIL:        {counts['FAIL']}")
    print(f"ERROR:       {counts['ERROR']}")
    print(f"SKIPPED:     {counts['SKIPPED']}")
    print("-" * 60)
    print(f"Total breaches: {total_breaches}")
    print(f"Output file: {output_path}")
    print("=" * 60 + "\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Compliance Rule Batch Runner")
    parser.add_argument("--fund-id", required=True, help="Fund ID to check (integer, e.g. 1)")
    parser.add_argument("--date", default=datetime.today().strftime("%Y-%m-%d"),
                        help="As-of date (YYYY-MM-DD)")
    parser.add_argument("--input", default=DEFAULT_RULES_PATH,
                        help=f"Path to input rules Excel (default: {DEFAULT_RULES_PATH})")
    parser.add_argument("--output", default=None,
                        help="Path to output Excel (default: compliance_results_<fund>_<date>.xlsx)")
    args = parser.parse_args()

    fund_id = args.fund_id
    as_of_date = args.date
    date_compact = as_of_date.replace("-", "")
    output_path = args.output or os.path.join(
        PROJECT_ROOT, f"compliance_results_{fund_id}_{date_compact}.xlsx"
    )

    print(f"Loading rules from: {args.input}")
    rules = load_rules(args.input)
    enabled_count = sum(1 for r in rules if str(r.get("enabled", "TRUE")).upper() == "TRUE")
    print(f"Loaded {len(rules)} rules ({enabled_count} enabled)")

    print(f"Running compliance checks for fund_id={fund_id}, as_of={as_of_date} ...")
    for r in rules:
        if str(r.get("enabled", "TRUE")).upper() == "TRUE":
            print(f"  Checking {r['rule_id']}: {r['rule_text'][:60]}")

    summary_rows, breach_rows = run_rules(rules, fund_id, as_of_date)
    write_output(output_path, summary_rows, breach_rows)
    print_summary(fund_id, as_of_date, summary_rows, output_path)


if __name__ == "__main__":
    main()
